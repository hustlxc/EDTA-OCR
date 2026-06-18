#!/usr/bin/env python3
"""Web viewer for EDTA-OCR merged database and captures."""

import base64
import sqlite3
import io
import os
import functools
import secrets
from flask import Flask, g, jsonify, request, render_template, session, redirect, url_for, Response, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

DB_PATH = "merged_embedded.db"
PER_PAGE = 40

# ── Box-Hole calculation ───────────────────────────────────────────────────
# Each specimen box holds 81 tubes in numbered holes (1–81).
# Tubes are placed sequentially in sort order of bullet number, starting from
# a user-chosen startBullet:
#   startRank  = rank of startBullet (1-based sort position)
#   offset     = rank - startRank
#   globalPos  = (firstBox - 1) * 81 + (firstHole - 1) + offset
#   box        = globalPos / 81 + 1
#   hole       = globalPos % 81 + 1
#
# Bullets with rank < startRank get box/hole = null (displayed as "NA").
#
# Set via env vars:  export FIRST_BOX=34  FIRST_HOLE=29  START_BULLET=2801
FIRST_BOX  = int(os.environ.get("FIRST_BOX",  "34"))
FIRST_HOLE = int(os.environ.get("FIRST_HOLE", "29"))
HOLES_PER_BOX = 81

# Runtime mutable box config (can be changed via API)
# startBullet: the bullet number at which box-hole calculation starts.
# Bullets with a smaller rank get box/hole = null (shown as "NA").
_box_config = {"firstBox": FIRST_BOX, "firstHole": FIRST_HOLE, "startBullet": "2801"}

# Cached record-id → sequential rank mapping (1-based).
# Each record gets a unique rank even when bullet numbers are duplicated.
_id_rank = {}          # int record_id → int rank
_rank_cache_db = None  # db path used to build the cache


def _build_id_rank(db):
    """Build id→rank lookup: all records sorted by bullet number, then by id.
    Duplicate bullet numbers get consecutive ranks."""
    global _id_rank, _rank_cache_db
    rows = db.execute(
        "SELECT id, 子弹头编号 FROM records WHERE 子弹头编号 != ''"
        " ORDER BY CAST(子弹头编号 AS INTEGER), id"
    ).fetchall()
    _id_rank = {}
    for i, r in enumerate(rows, start=1):
        _id_rank[r[0]] = i  # record_id → rank
    _rank_cache_db = DB_PATH


def record_rank(db, record_id):
    """Return the 1-based sequential rank of a record.  Builds cache on first call."""
    if _rank_cache_db != DB_PATH or not _id_rank:
        _build_id_rank(db)
    return _id_rank.get(record_id)


def min_bullet_number(db):
    """Return the smallest bullet number (for display only)."""
    row = db.execute(
        "SELECT MIN(CAST(子弹头编号 AS INTEGER)) FROM records WHERE 子弹头编号 != ''"
    ).fetchone()
    if row and row[0] is not None and row[0] > 0:
        return int(row[0])
    return None


def box_hole_for(rank, start_rank):
    """Calculate (box, hole) from the sorted rank of a bullet number.
    Returns None if rank < startRank (bullet is before the starting point),
    or if inputs are invalid."""
    if rank is None or start_rank is None:
        return None
    if rank < start_rank:
        return None  # NA — before the starting bullet
    first_box  = _box_config["firstBox"]
    first_hole = _box_config["firstHole"]
    if first_box <= 0 or not (1 <= first_hole <= HOLES_PER_BOX):
        return None
    offset = rank - start_rank
    global_pos = (first_box - 1) * HOLES_PER_BOX + (first_hole - 1) + offset
    if global_pos < 0:
        return None
    box = global_pos // HOLES_PER_BOX + 1
    hole = global_pos % HOLES_PER_BOX + 1
    return {"box": box, "hole": hole}

# Allowed sort columns — maps frontend key → SQL ORDER BY expression.
# Columns that store numbers as text use CAST(... AS INTEGER) for natural sort.
SORT_COLUMNS = {
    "id":     "id",
    "name":   "姓名",
    "gender": "性别",
    "age":    "年龄",
    "serial": "住院号",
    "bullet": "CAST(子弹头编号 AS INTEGER)",
    "time":   "采血时间",
    "dept":   "科室",
    "bed":    "床号",
    "saved":  "录入时间",
    "boxhole": "CAST(子弹头编号 AS INTEGER)",  # sort by bullet number for box-hole
}

# ── Authentication ─────────────────────────────────────────────────────────
# Set via environment variables, or use these defaults:
#   export WEB_USER=admin
#   export WEB_PASS=your-password-here
VALID_USER = os.environ.get("WEB_USER", "admin")
VALID_PASS = os.environ.get("WEB_PASS", "edta2026")

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", secrets.token_hex(32))


def login_required(f):
    """Decorator: supports session cookie AND HTTP Basic Auth.
    Redirects to /login only if neither is valid."""

    @functools.wraps(f)
    def decorated(*args, **kwargs):
        # 1. Session cookie
        if session.get("logged_in"):
            return f(*args, **kwargs)
        # 2. HTTP Basic Auth (for API clients)
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Basic "):
            try:
                creds = base64.b64decode(auth[6:]).decode("utf-8")
                user, _, pwd = creds.partition(":")
                if user == VALID_USER and pwd == VALID_PASS:
                    return f(*args, **kwargs)
            except Exception:
                pass
        # 3. Neither valid — redirect browser, 401 for API
        if request.path.startswith("/api/"):
            return jsonify({"error": "Unauthorized"}), 401
        return redirect(url_for("login", next=request.path))

    return decorated


# ── Database ────────────────────────────────────────────────────────────────


def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA busy_timeout=5000")
    return g.db


@app.teardown_appcontext
def close_db(_exception=None):
    db = g.pop("db", None)
    if db:
        db.close()


# ── Auth routes ─────────────────────────────────────────────────────────────


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        if (
            request.form.get("username") == VALID_USER
            and request.form.get("password") == VALID_PASS
        ):
            session["logged_in"] = True
            session.permanent = True
            next_url = request.args.get("next", "/")
            return redirect(next_url)
        error = "用户名或密码错误"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ── App routes ──────────────────────────────────────────────────────────────


@app.route("/")
@login_required
def index():
    return render_template("index.html")


@app.route("/api/departments")
@login_required
def api_departments():
    db = get_db()
    rows = db.execute(
        "SELECT 科室, COUNT(*) AS cnt FROM records WHERE 科室 != '' GROUP BY 科室 ORDER BY cnt DESC"
    ).fetchall()
    return jsonify([{"name": r["科室"], "count": r["cnt"]} for r in rows])


@app.route("/api/box-config", methods=["GET", "POST"])
@login_required
def api_box_config():
    global _box_config
    db = get_db()
    if request.method == "POST":
        data = request.get_json(silent=True) or {}
        if "firstBox" in data:
            _box_config["firstBox"] = max(1, int(data["firstBox"]))
        if "firstHole" in data:
            _box_config["firstHole"] = min(HOLES_PER_BOX, max(1, int(data["firstHole"])))
        if "startBullet" in data:
            val = data["startBullet"]
            if val is not None and str(val).strip():
                _box_config["startBullet"] = str(val).strip()
            else:
                _box_config["startBullet"] = None
    # Default startBullet to the smallest bullet if not set
    sb = _box_config["startBullet"]
    if sb is None:
        mb = min_bullet_number(db)
        sb = str(mb) if mb is not None else None
    # startRank = rank of the first record with startBullet
    start_rank = None
    if sb:
        first_id = db.execute(
            "SELECT MIN(id) FROM records WHERE 子弹头编号 = ?", (sb,)
        ).fetchone()
        if first_id and first_id[0]:
            start_rank = record_rank(db, first_id[0])
    return jsonify({
        "firstBox":   _box_config["firstBox"],
        "firstHole":  _box_config["firstHole"],
        "startBullet": sb,
        "startRank":  start_rank,
        "minBullet":  min_bullet_number(db),
    })


@app.route("/api/stats")
@login_required
def api_stats():
    db = get_db()
    total = db.execute("SELECT COUNT(*) FROM records").fetchone()[0]
    has_img = db.execute(
        "SELECT COUNT(*) FROM records WHERE 图片 IS NOT NULL"
    ).fetchone()[0]
    return jsonify({
        "total": total,
        "with_bullet": has_img,
        "minBullet": min_bullet_number(db),
        "firstBox":  _box_config["firstBox"],
        "firstHole": _box_config["firstHole"],
    })


@app.route("/api/records")
@login_required
def api_records():
    db = get_db()

    search = request.args.get("q", "").strip()
    dept = request.args.get("dept", "").strip()
    sort = request.args.get("sort", "id").strip()
    order = request.args.get("order", "asc").strip().upper()
    page = max(1, int(request.args.get("page", 1)))

    # Whitelist sort column; default to id
    sort_col = SORT_COLUMNS.get(sort, "id")
    # Whitelist order direction
    if order not in ("ASC", "DESC"):
        order = "ASC"

    conditions = []
    params = []

    if search:
        conditions.append(
            "(姓名 LIKE ? OR 住院号 LIKE ? OR 子弹头编号 LIKE ? OR 床号 LIKE ?)"
        )
        like = f"%{search}%"
        params.extend([like, like, like, like])

    if dept:
        conditions.append("科室 = ?")
        params.append(dept)

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    # count
    count = db.execute(
        f"SELECT COUNT(*) FROM records {where}", params
    ).fetchone()[0]

    total_pages = max(1, (count + PER_PAGE - 1) // PER_PAGE)
    page = min(page, total_pages)
    offset = (page - 1) * PER_PAGE

    # Select columns explicitly — exclude BLOB to avoid loading MBs per page
    rows = db.execute(
        f"SELECT id, 姓名, 性别, 年龄, 住院号, 子弹头编号, 采血时间, 科室, 床号, 原始OCR文本, 录入时间,"
        f" 图片 IS NOT NULL AS has_image"
        f" FROM records {where} ORDER BY {sort_col} {order} LIMIT ? OFFSET ?",
        params + [PER_PAGE, offset],
    ).fetchall()

    min_bullet = min_bullet_number(db)
    sb = _box_config.get("startBullet")
    if not sb:
        sb = str(min_bullet) if min_bullet is not None else None
    start_rank = None
    if sb:
        first_id = db.execute(
            "SELECT MIN(id) FROM records WHERE 子弹头编号 = ?", (sb,)
        ).fetchone()
        if first_id and first_id[0]:
            start_rank = record_rank(db, first_id[0])

    records = []
    for r in rows:
        bullet = r["子弹头编号"] or ""
        has_image = bool(r["has_image"])
        rank = record_rank(db, r["id"])
        bh = box_hole_for(rank, start_rank)
        records.append(
            {
                "id": r["id"],
                "name": r["姓名"] or "",
                "gender": r["性别"] or "",
                "age": r["年龄"] or "",
                "serial": r["住院号"] or "",
                "bullet": bullet,
                "time": r["采血时间"] or "",
                "dept": r["科室"] or "",
                "bed": r["床号"] or "",
                "ocr": r["原始OCR文本"] or "",
                "saved": r["录入时间"] or "",
                "has_image": has_image,
                "box": bh["box"] if bh else None,
                "hole": bh["hole"] if bh else None,
            }
        )

    return jsonify(
        {
            "records": records,
            "page": page,
            "total_pages": total_pages,
            "total": count,
            "sort": sort,
            "order": order,
        }
    )


@app.route("/api/export")
@login_required
def api_export():
    """Export current view to Excel (.xlsx).  Respects search/department filter."""
    db = get_db()

    search = request.args.get("q", "").strip()
    dept = request.args.get("dept", "").strip()
    sort = request.args.get("sort", "bullet").strip()
    order = request.args.get("order", "asc").strip().upper()

    sort_col = SORT_COLUMNS.get(sort, "CAST(子弹头编号 AS INTEGER)")
    if order not in ("ASC", "DESC"):
        order = "ASC"

    conditions = []
    params = []
    if search:
        conditions.append(
            "(姓名 LIKE ? OR 住院号 LIKE ? OR 子弹头编号 LIKE ? OR 床号 LIKE ?)"
        )
        like = f"%{search}%"
        params.extend([like, like, like, like])
    if dept:
        conditions.append("科室 = ?")
        params.append(dept)
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    # Fetch all matching records (no pagination) — exclude BLOB
    rows = db.execute(
        f"SELECT id, 姓名, 性别, 年龄, 住院号, 子弹头编号, 采血时间, 科室, 床号, 原始OCR文本, 录入时间"
        f" FROM records {where} ORDER BY {sort_col} {order}",
        params,
    ).fetchall()

    # Box-hole config
    sb = _box_config.get("startBullet")
    if not sb:
        mb = min_bullet_number(db)
        sb = str(mb) if mb is not None else None
    start_rank = None
    if sb:
        first_id = db.execute(
            "SELECT MIN(id) FROM records WHERE 子弹头编号 = ?", (sb,)
        ).fetchone()
        if first_id and first_id[0]:
            start_rank = record_rank(db, first_id[0])

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "EDTA-OCR 记录"

    # Header style
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["序号", "姓名", "性别", "年龄", "住院号", "子弹头编号", "采血时间", "科室", "床号", "盒·孔", "原始OCR文本", "录入时间"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for i, r in enumerate(rows):
        rank = record_rank(db, r["id"])
        bh = box_hole_for(rank, start_rank)
        boxhole = f"{bh['box']}-{bh['hole']}" if bh else "NA"
        values = [
            i + 1,
            r["姓名"] or "", r["性别"] or "", r["年龄"] or "",
            r["住院号"] or "", r["子弹头编号"] or "",
            r["采血时间"] or "", r["科室"] or "", r["床号"] or "",
            boxhole,
            r["原始OCR文本"] or "", r["录入时间"] or "",
        ]
        row_num = i + 2
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.border = thin_border
            if col != 1:
                cell.alignment = Alignment(horizontal="center" if col <= 10 else "left")

    # Column widths
    widths = [6, 10, 6, 6, 12, 14, 16, 26, 10, 10, 40, 20]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    # Write to memory and return
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "edta_ocr_export.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/record/<int:record_id>")
@login_required
def api_record(record_id):
    db = get_db()
    r = db.execute("SELECT * FROM records WHERE id = ?", (record_id,)).fetchone()
    if not r:
        return jsonify({"error": "Not found"}), 404
    bullet = r["子弹头编号"] or ""
    sb = _box_config.get("startBullet")
    if not sb:
        mb = min_bullet_number(db)
        sb = str(mb) if mb is not None else None
    start_rank = None
    if sb:
        first_id = db.execute(
            "SELECT MIN(id) FROM records WHERE 子弹头编号 = ?", (sb,)
        ).fetchone()
        if first_id and first_id[0]:
            start_rank = record_rank(db, first_id[0])
    bh = box_hole_for(record_rank(db, r["id"]), start_rank)
    has_img = bool(r["图片"]) if "图片" in r.keys() else False
    return jsonify(
        {
            "id": r["id"],
            "name": r["姓名"] or "",
            "gender": r["性别"] or "",
            "age": r["年龄"] or "",
            "serial": r["住院号"] or "",
            "bullet": bullet,
            "time": r["采血时间"] or "",
            "dept": r["科室"] or "",
            "bed": r["床号"] or "",
            "ocr": r["原始OCR文本"] or "",
            "saved": r["录入时间"] or "",
            "box": bh["box"] if bh else None,
            "hole": bh["hole"] if bh else None,
            "has_image": has_img,
        }
    )


@app.route("/api/records", methods=["POST"])
@login_required
def api_create_record():
    """Insert or replace a record.  'bullet' is the unique key (upsert)."""
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "Invalid JSON"}), 400
    bullet = (data.get("bullet") or "").strip()
    if not bullet:
        return jsonify({"error": "子弹头编号 is required"}), 400

    db = get_db()
    db.execute(
        "INSERT OR REPLACE INTO records (姓名, 性别, 年龄, 住院号, 子弹头编号, 采血时间, 科室, 床号, 原始OCR文本, 录入时间, 图片) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,"
        " COALESCE((SELECT 图片 FROM records WHERE 子弹头编号 = ?), NULL))",
        (
            data.get("name", "").strip() or None,
            data.get("gender", "").strip() or None,
            data.get("age", "").strip() or None,
            data.get("serial", "").strip() or None,
            bullet,
            data.get("time", "").strip() or None,
            data.get("dept", "").strip() or None,
            data.get("bed", "").strip() or None,
            data.get("ocr", "").strip() or None,
            data.get("saved") or None,
            bullet,
        ),
    )
    db.commit()
    # Invalidate rank cache so box-hole recalculates
    global _id_rank, _rank_cache_db
    _id_rank = {}
    _rank_cache_db = None

    # Return the new record
    r = db.execute("SELECT * FROM records WHERE 子弹头编号 = ?", (bullet,)).fetchone()
    return jsonify({"id": r["id"], "bullet": bullet, "name": r["姓名"] or ""}), 201


@app.route("/api/records/<int:record_id>", methods=["PUT"])
@login_required
def api_update_record(record_id):
    """Update a single record's metadata fields."""
    db = get_db()
    r = db.execute("SELECT id FROM records WHERE id = ?", (record_id,)).fetchone()
    if not r:
        return jsonify({"error": "Not found"}), 404

    data = request.get_json(silent=True) or {}
    fields = {
        "name": "姓名", "gender": "性别", "age": "年龄",
        "serial": "住院号", "bullet": "子弹头编号",
        "time": "采血时间", "dept": "科室", "bed": "床号",
        "ocr": "原始OCR文本",
    }
    for key, col in fields.items():
        if key in data:
            db.execute(
                f"UPDATE records SET {col} = ? WHERE id = ?",
                (data[key].strip() if data[key] else None, record_id),
            )
    db.commit()
    global _id_rank, _rank_cache_db
    _id_rank = {}
    _rank_cache_db = None
    return jsonify({"id": record_id})


@app.route("/api/records/<int:record_id>", methods=["DELETE"])
@login_required
def api_delete_record(record_id):
    """Delete a record and its embedded image."""
    db = get_db()
    r = db.execute("SELECT id FROM records WHERE id = ?", (record_id,)).fetchone()
    if not r:
        return jsonify({"error": "Not found"}), 404
    db.execute("DELETE FROM records WHERE id = ?", (record_id,))
    db.commit()
    global _id_rank, _rank_cache_db
    _id_rank = {}
    _rank_cache_db = None
    return jsonify({"deleted": record_id})


@app.route("/api/records/<int:record_id>/image", methods=["POST"])
@login_required
def api_upload_image(record_id):
    """Upload / replace the image for a record."""
    db = get_db()
    r = db.execute("SELECT id FROM records WHERE id = ?", (record_id,)).fetchone()
    if not r:
        return jsonify({"error": "Not found"}), 404
    img_data = request.get_data()
    if not img_data:
        return jsonify({"error": "No image data"}), 400
    db.execute("UPDATE records SET 图片 = ? WHERE id = ?", (img_data, record_id))
    db.commit()
    return jsonify({"id": record_id, "size": len(img_data)})


@app.route("/captures/<filename>")
@login_required
def serve_image(filename):
    # filename is "{record_id}.png" — look up BLOB by record ID, so duplicate
    # bullet numbers each serve their own image.
    if not filename.endswith(".png"):
        return "Bad filename", 400
    try:
        record_id = int(filename[:-4])
    except ValueError:
        return "Bad filename", 400
    db = get_db()
    row = db.execute(
        "SELECT 图片 FROM records WHERE id = ?", (record_id,)
    ).fetchone()
    if not row or not row[0]:
        return "Not found", 404
    return Response(row[0], mimetype="image/png")


if __name__ == "__main__":
    os.makedirs("templates", exist_ok=True)
    print(f"Database: {DB_PATH} ({os.path.getsize(DB_PATH)/1024/1024:.1f} MB)")
    print()
    print(f"  Username: {VALID_USER}")
    print(f"  Password: {VALID_PASS}")
    print(f"  (set WEB_USER / WEB_PASS env vars to change)")
    print(f"\nStarting server at http://localhost:8087")
    app.run(host="0.0.0.0", port=8087, debug=True)
